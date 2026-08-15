"""PGCB's daily NLDC workbook — the same reports as the BPDB PDFs, digitised.

Source: erp.powergrid.gov.bd publishes one .xlsx per day holding ten sheets.
Four of them (P1-P4) are the reports this project already reads out of BPDB's
PDFs, which means they double as a check on that parse; the rest is new:

  L-Curve    half-hourly generation, east grid / west grid / total
  En-Curve   half-hourly generation split fourteen ways by fuel, plus the
             half-hourly shortage — load-shedding at 30-minute resolution
  P1         system summary: peak generation and demand with their times,
             energy generated, energy unserved, maximum temperature, gas
             supplied, production cost per kWh, zone-by-fuel generation and
             the taka cost of each fuel
  P2         evening-peak generation for every plant
  P3         zone totals and the maximum load served by each grid sub-station
  P4         hourly generation, load-shed and demand
  Forecast   per-plant forecast availability against yesterday's actual

The report listing only ever shows the last thirty days, but the download
endpoint takes a numeric document id, and those run in sequence: the daily
workbooks occupy roughly 4431 upward, one per day from January 2025. Ids
below that, and the occasional id inside the range, belong to other documents
and simply 404 or return a workbook with no date, so the crawl walks the range
and keeps what parses.

  python scrape_pgcb_erp.py                # ids seen on the listing page
  python scrape_pgcb_erp.py --full         # walk the whole id range
  python scrape_pgcb_erp.py --from 4431 --to 4600

Output: raw/erp/*.csv and *.json, partitioned by year, plus manifest.json
        recording which ids have been fetched so a re-run resumes.
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import re
import sys
from collections import defaultdict

import openpyxl

from common import RAW, get, session, read_json, write_json, write_csv, read_csv

OUT = RAW / "erp"
LISTING = ("https://erp.powergrid.gov.bd/w/report/eyJpdiI6IldsU2ZQTGkvbkRnQU9F"
           "MjZ5UHhmeGc9PSIsInZhbHVlIjoiQzhONVl5ZGxRY3E3T3ZVNCtLZGt1Zz09Iiwi"
           "bWFjIjoiN2JiNTI5MzNhOWIxZDVjY2NkMmFlZWU4ZDU1N2I4OWZlYjNlZWM1ZGU4"
           "NzRiNWU4ZjQ3ZDc1ODRlMTk3MDc0YyIsInRhZyI6IiJ9/show_report")
DOWNLOAD = "https://erp.powergrid.gov.bd/web/files/download"

# The daily workbooks start here; ids below belong to other documents.
FIRST_ID = 4431
ID_CEILING = 20000          # a stop, not an expectation

# Sheet names drift over time: 'Volt' became 'Voltage', 'yesterdayGen' became
# 'GenLog'. Everything this module reads is matched case-insensitively against
# these, so a rename does not silently drop a day.
SHEET_ALIASES = {
    "p1": ("p1",), "p2": ("p2",), "p3": ("p3",), "p4": ("p4",),
    "lcurve": ("l-curve", "lcurve", "l curve"),
    "encurve": ("en-curve", "encurve", "en curve"),
    "forecast": ("forecast",),
}

# En-Curve's fuel columns, in the order the sheet writes them. Matched on the
# header text rather than position: a column inserted upstream would otherwise
# shift every fuel one place and go unnoticed.
FUEL_COLS = {
    "gas-public": "gas_public", "gas-pvt": "gas_pvt", "hvdc": "hvdc",
    "nepal": "nepal", "tripura": "tripura", "adani": "adani",
    "hydro": "hydro", "coal": "coal", "solar": "solar",
    "hfo-public": "hfo_public", "hfo-pvt": "hfo_pvt",
    "hsd-public": "hsd_public", "hsd-pvt": "hsd_pvt", "wind": "wind",
    "shortage": "shortage",
}
FUEL_FIELDS = list(dict.fromkeys(FUEL_COLS.values()))

HH_HEADER = ["datetime", "date", "time", "east", "west", "total"] + FUEL_FIELDS
HOURLY_HEADER = ["datetime", "date", "time", "generation", "loadshed", "demand"]
PLANT_HEADER = ["date", "plant", "producer", "units", "installed_mw",
                "present_mw", "peak_mw", "energy_kwh", "remarks"]
FORECAST_HEADER = ["date", "plant", "fuel", "producer", "units", "installed_mw",
                   "present_mw", "actual_day", "actual_evening",
                   "forecast_day", "forecast_evening", "effective_day",
                   "effective_evening", "remarks"]
SUBSTATION_HEADER = ["date", "substation", "load_mw", "time"]

ZONE_ROWS = {
    "dhaka": "dhaka", "chattogram": "chattogram", "chittagong": "chattogram",
    "cumilla": "cumilla", "comilla": "cumilla", "mymensingh": "mymensingh",
    "sylhet": "sylhet", "khulna": "khulna", "barishal": "barishal",
    "barisal": "barishal", "rajshahi": "rajshahi", "rangpur": "rangpur",
}


# ── small helpers ────────────────────────────────────────────────────────────

def cell(v):
    """A number, or None. Excel hands back ints, floats, strings and dashes."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s or s in {"-", "--", "N/A", "NA", "nil", "Nil"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def text(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def as_time(v):
    """'00:30', a datetime.time, or an Excel fraction of a day -> 'HH:MM'."""
    if isinstance(v, (dt.time, dt.datetime)):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)) and 0 <= v < 1:
        mins = round(v * 24 * 60)
        return f"{mins // 60:02d}:{mins % 60:02d}"
    s = text(v)
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    return f"{h:02d}:{mi:02d}" if h < 24 and mi < 60 else None


def as_date(v):
    """The workbooks write dates as text (11-08-2026) or as real dates."""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = text(v)
    m = re.search(r"(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})", s)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            return None
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


# Installed capacity is written as a unit expression -- "1*260", "2x120+1*50"
# -- and has no numeric column of its own, so the MW figure is worked out from
# the expression and the expression itself kept alongside it.
UNIT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[*x\u00d7X]\s*(\d+(?:\.\d+)?)")


def installed_mw(v):
    s = text(v)
    parts = UNIT_RE.findall(s)
    if parts:
        return round(sum(float(n) * float(c) for n, c in parts), 3)
    return cell(v)


def sheet(wb, key):
    """Find a sheet by any of its historical names."""
    want = SHEET_ALIASES[key]
    for name in wb.sheetnames:
        if name.strip().lower() in want:
            return wb[name]
    return None


def grid(ws, max_row=None, max_col=None):
    return [list(r) for r in ws.iter_rows(
        max_row=max_row or ws.max_row, max_col=max_col or ws.max_column,
        values_only=True)]


def find_date(rows, limit=12):
    """The 'Date :' label sits in a different column on every sheet."""
    for row in rows[:limit]:
        for j, v in enumerate(row):
            if text(v).lower().rstrip(" :") in {"date", "date :"}:
                for k in range(j + 1, min(j + 4, len(row))):
                    d = as_date(row[k])
                    if d:
                        return d
    for row in rows[:limit]:
        for v in row:
            d = as_date(v)
            if d:
                return d
    return None


# ── sheet parsers ────────────────────────────────────────────────────────────

def parse_lcurve(ws):
    """Half-hourly east / west / total generation."""
    rows = grid(ws)
    date = find_date(rows, limit=4)
    hdr = next((i for i, r in enumerate(rows[:6])
                if text(r[0]).lower() == "time"), None)
    if hdr is None or not date:
        return date, []
    out = []
    for r in rows[hdr + 1:]:
        t = as_time(r[0])
        if not t:
            continue
        east, west, total = cell(r[1]), cell(r[2]), cell(r[3])
        if total is None and east is None:
            continue
        out.append({"time": t, "east": east, "west": west, "total": total})
    return date, out


def parse_encurve(ws):
    """Half-hourly generation by fuel, and the half-hourly shortage."""
    rows = grid(ws)
    date = find_date(rows, limit=4)
    hdr = next((i for i, r in enumerate(rows[:6])
                if text(r[0]).lower() == "time"), None)
    if hdr is None or not date:
        return date, []
    cols = {}
    for j, v in enumerate(rows[hdr]):
        key = FUEL_COLS.get(text(v).lower())
        if key and key not in cols:
            cols[key] = j
    out = []
    for r in rows[hdr + 1:]:
        t = as_time(r[0])
        if not t:
            continue
        rec = {"time": t}
        for key, j in cols.items():
            rec[key] = cell(r[j]) if j < len(r) else None
        if any(rec.get(k) is not None for k in FUEL_FIELDS):
            out.append(rec)
    return date, out


def parse_p4(ws):
    """Hourly generation, load-shed and demand."""
    rows = grid(ws)
    date = find_date(rows)
    if not date:
        return None, []
    hdr = None
    for i, r in enumerate(rows[:20]):
        labels = [text(v).lower() for v in r]
        if "hour" in labels and any("generation" in x for x in labels):
            hdr = i
            break
    if hdr is None:
        return date, []
    cols = {}
    for j, v in enumerate(rows[hdr]):
        lab = text(v).lower()
        if lab == "hour":
            cols["hour"] = j
        elif "generation" in lab:
            cols["generation"] = j
        elif "load" in lab and "shed" in lab:
            cols["loadshed"] = j
        elif "demand" in lab:
            cols["demand"] = j
    if "hour" not in cols:
        return date, []
    out = []
    for r in rows[hdr + 1:]:
        t = as_time(r[cols["hour"]]) if cols["hour"] < len(r) else None
        if not t:
            continue
        rec = {"time": t}
        for k in ("generation", "loadshed", "demand"):
            j = cols.get(k)
            rec[k] = cell(r[j]) if j is not None and j < len(r) else None
        if rec["generation"] is not None or rec["demand"] is not None:
            out.append(rec)
    return date, out


# P1's left column is a label and the number sits a few cells to its right;
# the sheet has been re-laid-out at least once, so each figure is found by
# matching its label rather than by address.
P1_FIGURES = {
    "day_peak_generation": r"day\s*peak\s*generation",
    "day_peak_demand": r"day\s*peak\s*demand",
    "evening_peak_generation": r"evening\s*peak\s*generation",
    "evening_peak_demand": r"evening\s*peak\s*demand",
    "min_generation": r"minimum\s*generation",
    "max_generation": r"maximum\s*generation",
    "energy_generated": r"energy\s*generated",
    "energy_unserved": r"energy\s*unserve",     # the sheet spells it 'Unserverd'
    "energy_demand": r"energy\s*demand",
    "max_temperature": r"maximum\s*temperature",
    "gas_supplied": r"total\s*gas\s*supplied",
    "unit_cost": r"production\s*cost\s*per",
}
P1_TIMED = {"day_peak_generation", "day_peak_demand", "evening_peak_generation",
            "evening_peak_demand", "min_generation", "max_generation"}
P1_FUELS = ["gas", "coal", "hfo", "hsd", "hydro", "solar", "import", "wind"]


def parse_p1(ws):
    """System summary: headline figures, zone-by-fuel energy, cost by fuel."""
    rows = grid(ws)
    date = find_date(rows)
    if not date:
        return None, {}
    out = {}
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            lab = text(v).lower()
            if not lab:
                continue
            for key, pat in P1_FIGURES.items():
                if key in out or not re.search(pat, lab):
                    continue
                for k in range(j + 1, min(j + 6, len(r))):
                    n = cell(r[k])
                    if n is not None:
                        out[key] = n
                        if key in P1_TIMED:
                            for m in range(k + 1, min(k + 4, len(r))):
                                t = as_time(text(r[m]).replace("Hr.", "").strip())
                                if t:
                                    out[key + "_time"] = t
                                    break
                        break

    # zone-by-fuel generation (MkWh): a header row of fuel names, then one row
    # per zone and a Total row that this keeps for the reconciliation check
    zone_gen, zone_hdr = {}, None
    for i, r in enumerate(rows):
        labels = [text(v).lower() for v in r]
        if "gas" in labels and "coal" in labels and "total" in labels:
            zone_hdr = (i, {text(v).lower(): j for j, v in enumerate(r) if text(v)})
            break
    if zone_hdr:
        i0, cols = zone_hdr
        for r in rows[i0 + 1:i0 + 14]:
            name = next((text(v) for v in r[:4] if text(v)), "")
            key = ZONE_ROWS.get(name.lower().replace(" zone", "").strip())
            if key is None and name.lower().strip() != "total":
                continue
            key = key or "total"
            rec = {}
            for fuel in P1_FUELS + ["total"]:
                j = cols.get(fuel)
                if j is not None and j < len(r):
                    rec[fuel] = cell(r[j])
            zone_gen[key] = rec
    if zone_gen:
        out["zone_generation"] = zone_gen

    # 'Production Cost (Tk.)' — fuel / taka pairs, laid out in two columns
    cost = {}
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            if text(v).lower() != "fuel":
                continue
            for rr in rows[i + 1:i + 12]:
                if j >= len(rr):
                    continue
                fuel = text(rr[j]).lower()
                if fuel in P1_FUELS and j + 1 < len(rr):
                    n = cell(rr[j + 1])
                    if n is not None:
                        cost[fuel] = n
    if cost:
        out["cost_by_fuel_tk"] = cost
    return date, out


def parse_p3(ws):
    """Zone totals at evening peak, and each grid sub-station's maximum load."""
    rows = grid(ws)
    date = find_date(rows)
    if not date:
        return None, {}, []
    zones, totals = {}, {}
    for r in rows[:16]:
        for j, v in enumerate(r):
            lab = text(v).lower()
            if not lab:
                continue
            m = re.search(r"\b(?:i+v?|v i*|\d+)\)\s*([a-z]+)\s*area", lab)
            key = ZONE_ROWS.get(m.group(1)) if m else None
            if key:
                for k in range(j + 1, min(j + 5, len(r))):
                    n = cell(r[k])
                    if n is not None:
                        zones[key] = n
                        break
            elif "entire eastern" in lab or "entire western" in lab:
                side = "east" if "eastern" in lab else "west"
                for k in range(j + 1, min(j + 6, len(r))):
                    n = cell(r[k])
                    if n is not None:
                        totals[side] = n
                        break
            elif "system total demand" in lab:
                for k in range(j + 1, min(j + 8, len(r))):
                    n = cell(r[k])
                    if n is not None:
                        totals["system"] = n
                        break

    # The sub-station table repeats (Sl, name, load, time) across the page.
    hdr = None
    for i, r in enumerate(rows):
        labels = [text(v).lower() for v in r]
        if labels.count("sub-station") >= 1 and any("load" in x for x in labels):
            hdr = i
            break
    subs = []
    if hdr is not None:
        starts = [j for j, v in enumerate(rows[hdr])
                  if text(v).lower() == "sub-station"]
        for r in rows[hdr + 1:]:
            for j in starts:
                if j + 2 >= len(r):
                    continue
                name = text(r[j])
                load = cell(r[j + 1])
                if not name or load is None:
                    continue
                subs.append({"substation": name, "load_mw": load,
                             "time": as_time(r[j + 2]) or ""})
    return date, {"zones": zones, "totals": totals}, subs


def parse_p2(ws):
    """Evening-peak generation and day-long energy, plant by plant."""
    rows = grid(ws)
    date = find_date(rows)
    if not date:
        return None, []
    hdr = None
    for i, r in enumerate(rows[:20]):
        labels = [text(v).lower() for v in r]
        if any("name of the" in x for x in labels) and \
           any("peak hour" in x for x in labels):
            hdr = i
            break
    if hdr is None:
        return date, []
    cols = {}
    for j, v in enumerate(rows[hdr]):
        lab = text(v).lower()
        if "name of the" in lab:
            cols["plant"] = j
        elif "producer" in lab:
            cols["producer"] = j
        elif "installed" in lab:
            cols["installed_mw"] = j
        elif "present" in lab:
            cols["present_mw"] = j
        elif "peak hour" in lab:
            cols["peak_mw"] = j
        elif "energy generated" in lab:
            cols["energy_kwh"] = j
        elif "remarks" in lab:
            cols["remarks"] = j
    out = []
    for r in rows[hdr + 1:]:
        j = cols.get("plant")
        name = text(r[j]) if j is not None and j < len(r) else ""
        if not name or name.lower() in {"total", "grand total"}:
            continue
        rec = {"plant": name}
        for k in ("producer", "remarks"):
            jj = cols.get(k)
            rec[k] = text(r[jj]) if jj is not None and jj < len(r) else ""
        jj = cols.get("installed_mw")
        rec["units"] = text(r[jj]) if jj is not None and jj < len(r) else ""
        rec["installed_mw"] = installed_mw(rec["units"])
        for k in ("present_mw", "peak_mw", "energy_kwh"):
            jj = cols.get(k)
            rec[k] = cell(r[jj]) if jj is not None and jj < len(r) else None
        if rec["present_mw"] is None and rec["peak_mw"] is None:
            continue
        out.append(rec)
    return date, out


def parse_forecast(ws):
    """Per-plant forecast availability against yesterday's actual generation.

    The header spans three rows: a label row, a date row and a Day/Ev. Peak
    row, so the two sub-columns under each heading are taken by position from
    the heading's own column.
    """
    rows = grid(ws)
    hdr = None
    for i, r in enumerate(rows[:20]):
        labels = [text(v).lower() for v in r]
        if any("name of the" in x for x in labels) and \
           any("forecasted available" in x for x in labels):
            hdr = i
            break
    if hdr is None:
        return None, []
    # the forecast sheet is dated for the day being forecast
    date = find_date(rows, limit=hdr)
    cols = {}
    for j, v in enumerate(rows[hdr]):
        lab = text(v).lower()
        if "name of the" in lab:
            cols["plant"] = j
        elif lab == "fuel":
            cols["fuel"] = j
        elif "producer" in lab:
            cols["producer"] = j
        elif "installed" in lab:
            cols["installed_mw"] = j
        elif "present capacity" in lab:
            cols["present_mw"] = j
        elif "actual generation" in lab:
            cols["actual"] = j
        elif "forecasted available" in lab:
            cols["forecast"] = j
        elif "effective available" in lab:
            cols["effective"] = j
        elif "remarks" in lab:
            cols["remarks"] = j
    if "plant" not in cols:
        return date, []
    out = []
    for r in rows[hdr + 1:]:
        j = cols["plant"]
        name = text(r[j]) if j < len(r) else ""
        if not name or name.lower() in {"total", "grand total"}:
            continue
        rec = {"plant": name}
        for k in ("fuel", "producer", "remarks"):
            jj = cols.get(k)
            rec[k] = text(r[jj]) if jj is not None and jj < len(r) else ""
        jj = cols.get("installed_mw")
        rec["units"] = text(r[jj]) if jj is not None and jj < len(r) else ""
        rec["installed_mw"] = installed_mw(rec["units"])
        jj = cols.get("present_mw")
        rec["present_mw"] = cell(r[jj]) if jj is not None and jj < len(r) else None
        for base, out_a, out_b in (("actual", "actual_day", "actual_evening"),
                                   ("forecast", "forecast_day", "forecast_evening"),
                                   ("effective", "effective_day", "effective_evening")):
            jj = cols.get(base)
            rec[out_a] = cell(r[jj]) if jj is not None and jj < len(r) else None
            rec[out_b] = cell(r[jj + 1]) if jj is not None and jj + 1 < len(r) else None
        if rec["present_mw"] is None and rec["forecast_day"] is None:
            continue
        out.append(rec)
    return date, out


# ── one workbook ─────────────────────────────────────────────────────────────

def parse_workbook(blob):
    """Everything one day's workbook has to say, keyed by its own dates.

    Sheets carry two different dates: P1-P4 and the curves describe yesterday,
    while Forecast describes the day the file was published. Each is recorded
    under the date its own sheet states, never under the id or the file title,
    which have both been wrong upstream before.
    """
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=False)
    day = {"sheets": wb.sheetnames}

    ws = sheet(wb, "lcurve")
    if ws is not None:
        d, rows = parse_lcurve(ws)
        if d and rows:
            day["lcurve"] = {"date": d, "rows": rows}

    ws = sheet(wb, "encurve")
    if ws is not None:
        d, rows = parse_encurve(ws)
        if d and rows:
            day["encurve"] = {"date": d, "rows": rows}

    ws = sheet(wb, "p4")
    if ws is not None:
        d, rows = parse_p4(ws)
        if d and rows:
            day["hourly"] = {"date": d, "rows": rows}

    ws = sheet(wb, "p1")
    if ws is not None:
        d, rec = parse_p1(ws)
        if d and rec:
            day["summary"] = {"date": d, "rec": rec}

    ws = sheet(wb, "p3")
    if ws is not None:
        d, zones, subs = parse_p3(ws)
        if d and (zones["zones"] or subs):
            day["zones"] = {"date": d, "rec": zones}
            if subs:
                day["substations"] = {"date": d, "rows": subs}

    ws = sheet(wb, "p2")
    if ws is not None:
        d, rows = parse_p2(ws)
        if d and rows:
            day["plants"] = {"date": d, "rows": rows}

    ws = sheet(wb, "forecast")
    if ws is not None:
        d, rows = parse_forecast(ws)
        if d and rows:
            day["forecast"] = {"date": d, "rows": rows}

    wb.close()
    return day


def fetch(sess, doc_id):
    r = get(sess, DOWNLOAD, tries=3, timeout=90,
            params={"location": f"erp/web/report_docs/{doc_id}.xlsx",
                    "title": "Daily_Report"})
    if r is None:
        return None
    if not r.content[:2] == b"PK":
        return None
    return r.content


# ── accumulation and output ──────────────────────────────────────────────────

class Store:
    """Collects rows in memory, then writes one file per year.

    Every table is keyed so a day fetched twice replaces rather than
    duplicates: ids have overlapped before when a report was re-published.
    """

    def __init__(self):
        self.hh = {}            # (date, time) -> row
        self.hourly = {}
        self.summary = {}
        self.zones = {}
        self.subs = defaultdict(dict)     # date -> name -> row
        self.plants = defaultdict(dict)
        self.forecast = defaultdict(dict)

    def load(self):
        """Read back what previous runs wrote so a partial crawl resumes."""
        for path in sorted(OUT.glob("halfhourly_*.csv")):
            for r in read_csv(path):
                self.hh[(r["date"], r["time"])] = r
        for path in sorted(OUT.glob("hourly_*.csv")):
            for r in read_csv(path):
                self.hourly[(r["date"], r["time"])] = r
        for path in sorted(OUT.glob("summary_*.json")):
            self.summary.update(read_json(path, {}))
        for path in sorted(OUT.glob("zones_*.json")):
            self.zones.update(read_json(path, {}))
        for path in sorted(OUT.glob("substations_*.csv")):
            for r in read_csv(path):
                self.subs[r["date"]][r["substation"]] = r
        for path in sorted(OUT.glob("plants_*.csv")):
            for r in read_csv(path):
                self.plants[r["date"]][r["plant"]] = r
        for path in sorted(OUT.glob("forecast_*.csv")):
            for r in read_csv(path):
                self.forecast[r["date"]][r["plant"]] = r

    def add(self, day):
        lc = day.get("lcurve")
        ec = day.get("encurve")
        if lc or ec:
            date = (lc or ec)["date"]
            merged = {}
            for r in (lc["rows"] if lc else []):
                merged.setdefault(r["time"], {})["lc"] = r
            for r in (ec["rows"] if ec else []):
                merged.setdefault(r["time"], {})["ec"] = r
            # A curve sheet may carry its own date; keep them apart if so.
            ec_date = ec["date"] if ec else None
            for t, parts in merged.items():
                d = lc["date"] if "lc" in parts else ec_date
                row = {"datetime": f"{d} {t}", "date": d, "time": t}
                lcr = parts.get("lc", {})
                row["east"], row["west"] = lcr.get("east"), lcr.get("west")
                row["total"] = lcr.get("total")
                ecr = parts.get("ec", {})
                for f in FUEL_FIELDS:
                    row[f] = ecr.get(f)
                self.hh[(d, t)] = row

        h = day.get("hourly")
        if h:
            for r in h["rows"]:
                self.hourly[(h["date"], r["time"])] = {
                    "datetime": f"{h['date']} {r['time']}", "date": h["date"],
                    "time": r["time"], "generation": r["generation"],
                    "loadshed": r["loadshed"], "demand": r["demand"]}

        s = day.get("summary")
        if s:
            self.summary[s["date"]] = s["rec"]
        z = day.get("zones")
        if z:
            self.zones[z["date"]] = z["rec"]
        sb = day.get("substations")
        if sb:
            for r in sb["rows"]:
                self.subs[sb["date"]][r["substation"]] = {
                    "date": sb["date"], **r}
        p = day.get("plants")
        if p:
            for r in p["rows"]:
                self.plants[p["date"]][r["plant"]] = {"date": p["date"], **r}
        f = day.get("forecast")
        if f:
            for r in f["rows"]:
                self.forecast[f["date"]][r["plant"]] = {"date": f["date"], **r}

    def write(self):
        def by_year(rows, keyfn):
            out = defaultdict(list)
            for r in rows:
                out[keyfn(r)[:4]].append(r)
            return out

        # write_csv takes rows of values, not mappings: handing it dicts writes
        # their keys on every line.
        def rows_of(records, header):
            return [[r.get(c) for c in header] for r in records]

        for year, rows in by_year(self.hh.values(), lambda r: r["date"]).items():
            rows.sort(key=lambda r: (r["date"], r["time"]))
            write_csv(OUT / f"halfhourly_{year}.csv", rows_of(rows, HH_HEADER),
                      HH_HEADER)
        for year, rows in by_year(self.hourly.values(), lambda r: r["date"]).items():
            rows.sort(key=lambda r: (r["date"], r["time"]))
            write_csv(OUT / f"hourly_{year}.csv", rows_of(rows, HOURLY_HEADER),
                      HOURLY_HEADER)

        for name, table in (("summary", self.summary), ("zones", self.zones)):
            grouped = defaultdict(dict)
            for d, rec in table.items():
                grouped[d[:4]][d] = rec
            for year, rec in grouped.items():
                write_json(OUT / f"{name}_{year}.json",
                           dict(sorted(rec.items())), indent=1)

        for name, table, header in (
                ("substations", self.subs, SUBSTATION_HEADER),
                ("plants", self.plants, PLANT_HEADER),
                ("forecast", self.forecast, FORECAST_HEADER)):
            grouped = defaultdict(list)
            for d, byname in table.items():
                grouped[d[:4]].extend(byname.values())
            for year, rows in grouped.items():
                rows.sort(key=lambda r: (r["date"], r.get("substation") or
                                         r.get("plant") or ""))
                write_csv(OUT / f"{name}_{year}.csv",
                          rows_of(rows, header), header)


def listing_ids(sess):
    """The ids on the public listing page — what an hourly run needs."""
    r = get(sess, LISTING, timeout=60)
    if r is None:
        return []
    return sorted({int(m) for m in
                   re.findall(r"report_docs%2F(\d+)\.xlsx", r.text)})


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--full", action="store_true",
                    help="walk the whole id range, not just the listing page")
    ap.add_argument("--from", dest="lo", type=int, default=None)
    ap.add_argument("--to", dest="hi", type=int, default=None)
    ap.add_argument("--retry-missing", action="store_true",
                    help="try ids previously recorded as absent")
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)
    manifest = read_json(OUT / "manifest.json", {}) or {}
    done = manifest.setdefault("ids", {})
    missing = set(manifest.setdefault("missing", []))

    sess = session()
    listed = set()
    if args.lo or args.hi or args.full:
        lo = args.lo or FIRST_ID
        hi = args.hi or (max(int(k) for k in done) + 400 if done else FIRST_ID + 800)
        ids = list(range(lo, min(hi, ID_CEILING) + 1))
    else:
        listed = set(listing_ids(sess))
        if not listed:
            print("[erp] listing unreachable", file=sys.stderr)
            return 1
        # a few ids past the newest, in case the listing lags
        ids = sorted(listed) + list(range(max(listed) + 1, max(listed) + 4))

    # An id that 404s only because tomorrow's report does not exist yet must
    # not be blacklisted for good: a backfill that runs past the frontier
    # would otherwise freeze the collection, and the scraper would go on
    # reporting "nothing new" while reports piled up behind it. So the
    # missing set is only trusted below the newest id already held, and never
    # for an id the listing page is currently advertising.
    frontier = max((int(k) for k in done), default=FIRST_ID)

    def skip(i):
        if str(i) in done:
            return True
        if args.retry_missing or i in listed or i >= frontier:
            return False
        return i in missing

    todo = [i for i in ids if not skip(i)]
    if not todo:
        print("[erp] nothing new")
        return 0

    store = Store()
    store.load()
    print(f"[erp] {len(todo)} ids to try "
          f"({todo[0]}..{todo[-1]}); {len(done)} already held")

    ok = gone = bad = 0
    for n, doc_id in enumerate(todo, 1):
        blob = fetch(sess, doc_id)
        if blob is None:
            missing.add(doc_id)
            gone += 1
        else:
            try:
                day = parse_workbook(blob)
            except Exception as e:                      # noqa: BLE001
                print(f"[erp] {doc_id}: unreadable ({type(e).__name__}: {e})",
                      file=sys.stderr)
                bad += 1
                missing.add(doc_id)
                day = None
            if day:
                dates = {k: v["date"] for k, v in day.items()
                         if isinstance(v, dict) and v.get("date")}
                if not dates:
                    missing.add(doc_id)
                    bad += 1
                else:
                    store.add(day)
                    missing.discard(doc_id)
                    done[str(doc_id)] = {
                        "date": dates.get("hourly") or dates.get("lcurve")
                        or sorted(dates.values())[0],
                        "bytes": len(blob), "parts": sorted(dates)}
                    ok += 1
        if n % 25 == 0 or n == len(todo):
            store.write()
            manifest["missing"] = sorted(missing)
            write_json(OUT / "manifest.json", manifest, indent=1)
            print(f"[erp] {n}/{len(todo)}  kept {ok}  absent {gone}  "
                  f"unusable {bad}", flush=True)

    store.write()
    manifest["missing"] = sorted(missing)
    write_json(OUT / "manifest.json", manifest, indent=1)
    days = {v["date"] for v in done.values()}
    print(f"[erp] {ok} workbooks added; {len(days)} days held "
          f"({min(days) if days else '-'} to {max(days) if days else '-'})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
